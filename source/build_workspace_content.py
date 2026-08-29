#!/usr/bin/env python3
"""
Converts the authored 90-day workbook into the Naleli Workspace content
package the app loads at runtime.

    source/Naleli_Task_Based_Digital_Foundation_90_Day_Workbook.xlsx
        -> content/digital-foundation/workspace-content.json

The workbook is the curriculum's source of truth; this script is the only
thing that should ever write workspace-content.json, so the two can never
drift by hand-editing. Re-run it after any workbook change:

    python3 source/build_workspace_content.py

MAPPING (workbook -> app model)
    Stage         -> Phase        (4, matching course.json's stages)
    Module x Stage-> Workstream   (named for the module's Main Skill)
    Day           -> Task         (one task per day)
    Day's tasks   -> Sub-steps    (Learn / Practice / Work Mission / Review)

A module that runs across a stage boundary becomes two workstreams, one in
each phase, because a phase is defined by its day range — the later one is
labelled "(continued)" so the split is visible rather than silently
reordering the curriculum.

DERIVED, NOT FROM SOURCE (documented so nothing here reads as authored):
    tier             - the workbook asks for every task to be completed, so
                       days are REQUIRED; the capstone/portfolio days (80-90),
                       which the PORTFOLIO sheet lists with their own distinct
                       deliverables, are ASSESSMENT.
    estimatedMinutes - not in the workbook. 45 min for a standard day split
                       across its sub-steps, 90 for a capstone day.
    unlock order     - course.json sets SEQUENTIAL_UNLOCK; the app derives
                       "day N needs day N-1" itself rather than storing 89
                       prerequisite rows here.
"""

import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "source" / "Naleli_Task_Based_Digital_Foundation_90_Day_Workbook.xlsx"
OUT = ROOT / "content" / "digital-foundation" / "workspace-content.json"

# Workbook stage label -> the stageId used in course.json.
STAGE_IDS = {
    "Learn the Role": "stage-1",
    "Do the Work": "stage-2",
    "Operate Independently": "stage-3",
    "Capstone & Portfolio": "stage-4",
}

CAPSTONE_FIRST_DAY = 80
STANDARD_DAY_MINUTES = 45
CAPSTONE_DAY_MINUTES = 90

# The roadmap's Main Skill column is free text, so title-casing it alone
# yields "Generative Ai", and module 20 is labelled two different ways
# either side of the stage-3/4 boundary. Normalise to one display name per
# skill so the same workstream never appears under two spellings.
SKILL_DISPLAY_NAMES = {
    "foundations": "Foundations",
    "practice": "Keyboarding Practice",
    "digital citizenship": "Digital Citizenship",
    "online research": "Online Research",
    "safety": "Online Safety",
    "generative ai": "Generative AI",
    "computer operations": "Computer Operations",
    "software": "Software",
    "word processing": "Word Processing",
    "spreadsheets": "Spreadsheets",
    "presentations": "Presentations",
    "communication": "Communication",
    "hardware": "Hardware",
    "networking": "Networking",
    "cloud": "Cloud",
    "security": "Security",
    "data": "Databases",
    "coding": "Coding",
    "planning and innovation": "Planning and Innovation",
    "career/capstone": "Career Readiness",
    "career readiness / capstone": "Career Readiness",
}


def display_skill(skill):
    return SKILL_DISPLAY_NAMES.get(clean(skill).lower(), clean(skill).title())


def clean(value):
    """Workbook cells carry hard line breaks and doubled spacing."""
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def slug(text, limit=48):
    s = re.sub(r"[^a-z0-9]+", "-", clean(text).lower()).strip("-")
    return s[:limit].strip("-")


def read_roadmap(wb):
    """Day -> {stage, module, lesson, page, skill} for all 90 days."""
    rows = OrderedDict()
    for r in wb["ROADMAP"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        day = int(r[0])
        rows[day] = {
            "day": day,
            "stage": clean(r[1]),
            "modules": [m for m in (clean(x) for x in str(r[2]).split(",")) if m],
            "lesson": clean(r[3]),
            "page": clean(r[4]),
            "skills": [s for s in (clean(x) for x in str(r[5]).split(",")) if s],
        }
    return rows


def read_day_sheet(wb, day):
    """The per-day tab: header fields, its 4 tasks, and its review questions."""
    ws = wb[f"DAY {day:02d}"]
    grid = [[clean(c) for c in row] for row in ws.iter_rows(values_only=True)]

    def header(label):
        for row in grid:
            if row and row[0].lower() == label.lower():
                return next((c for c in row[1:] if c), "")
        return ""

    tasks, questions, in_tasks, in_questions = [], [], False, False
    for row in grid:
        first = row[0] if row else ""
        if first == "TASKS":
            in_tasks, in_questions = True, False
            continue
        if first == "SOURCE LEARNING QUESTIONS":
            in_tasks, in_questions = False, True
            continue
        if in_tasks and first.startswith(("Task", "Review")) and first != "Step":
            tasks.append(
                {
                    "step": first,
                    "title": row[1] if len(row) > 1 else "",
                    "instructions": row[2] if len(row) > 2 else "",
                    "evidence": row[3] if len(row) > 3 else "",
                }
            )
        elif in_questions and first == "Question":
            q = row[1] if len(row) > 1 else ""
            if not q:
                continue
            # A long question wraps onto a second "Question" row in the
            # workbook (28 of the 359 do). A fragment starting lower-case is
            # the tail of the previous question, not a new one.
            if questions and q[:1].islower():
                questions[-1] = f"{questions[-1]} {q}"
            else:
                questions.append(q)

    return {
        "stage": header("Stage"),
        "sourceLessons": header("Source lesson(s)"),
        "sourcePages": header("Source page(s)"),
        "learningFocus": header("Learning focus"),
        "tasks": tasks,
        "questions": questions,
    }


def criteria_for(sub_steps, deliverable):
    """The rubric AssessmentEngine actually applies, stated in the learner's
    words. These are DERIVED, not workbook text: the workbook's SOURCE
    LEARNING QUESTIONS are review questions the learner answers, not things
    an assessor ticks, so they go to reviewQuestions instead. Order matters
    — AssessmentEngine checks the last criterion against the evidence's file
    type, so the deliverable one stays last."""
    return [
        f"All {len(sub_steps)} steps of the day are complete",
        "Evidence of the work is attached",
        f"The evidence is the deliverable asked for: {deliverable}",
    ]


def first_work_task(sheet):
    """The day's first task that is not the closing Review — on capstone
    days this is the day's actual piece of work, and the only place its
    real name appears."""
    empty = {"step": "", "title": "", "instructions": "", "evidence": ""}
    return next(
        (
            t
            for t in sheet["tasks"]
            if not (t["title"] or t["step"]).lower().startswith("review")
        ),
        empty,
    )


def split_minutes(total, count):
    """Whole minutes per sub-step that always re-add to the task total."""
    if count == 0:
        return []
    base, extra = divmod(total, count)
    return [base + (1 if i < extra else 0) for i in range(count)]


def build():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    roadmap = read_roadmap(wb)

    # Workstreams keyed by (module, stage) so a module crossing a stage
    # boundary lands in both phases rather than dragging days into the
    # wrong one.
    workstreams = OrderedDict()
    seen_modules = set()

    for day, meta in roadmap.items():
        sheet = read_day_sheet(wb, day)
        stage = sheet["stage"] or meta["stage"]
        module = meta["modules"][0] if meta["modules"] else "M?"
        skill = display_skill(meta["skills"][0] if meta["skills"] else sheet["learningFocus"])
        is_capstone = day >= CAPSTONE_FIRST_DAY

        key = (module, stage)
        if key not in workstreams:
            continued = module in seen_modules
            seen_modules.add(module)
            workstreams[key] = {
                "workstreamId": f"ws-{slug(module)}-{slug(stage, 20)}",
                "name": skill + (" (continued)" if continued else ""),
                "stageId": STAGE_IDS[stage],
                "moduleCode": module,
                "tasks": [],
            }

        total = CAPSTONE_DAY_MINUTES if is_capstone else STANDARD_DAY_MINUTES
        per_step = split_minutes(total, len(sheet["tasks"]))
        sub_steps = [
            {
                "subStepId": f"day-{day:02d}-step-{i + 1}",
                "title": t["title"] or t["step"],
                "estimatedMinutes": per_step[i],
                "instructions": t["instructions"],
                "evidence": t["evidence"],
            }
            for i, t in enumerate(sheet["tasks"])
        ]

        # The workbook splits a day's work into Learn / Practice / Work
        # Mission / Review. Those map onto the Task Workspace's own
        # sections, so pull each one's instructions into the section it
        # belongs to instead of flattening them into one wall of text.
        def instructions_for(*prefixes):
            for t in sheet["tasks"]:
                label = (t["title"] or t["step"]).lower()
                if any(label.startswith(p) for p in prefixes):
                    return t["instructions"]
            return ""

        def evidence_for(*prefixes):
            for t in sheet["tasks"]:
                label = (t["title"] or t["step"]).lower()
                if any(label.startswith(p) for p in prefixes):
                    return t["evidence"]
            return ""

        learn = instructions_for("learn")
        practice = instructions_for("practice", "practise")
        mission = instructions_for("work mission", "portfolio", "capstone")
        if not mission:
            # Capstone days (80-90) name their tasks for the artefact being
            # produced — "Role-to-Skill Map", "Build — Part 1", "Present" —
            # rather than the standard Learn/Practice/Work Mission labels,
            # so fall back to the day's first non-review task.
            mission = first_work_task(sheet)["instructions"]
        # The day's deliverable is what it finally produces — the Work
        # Mission's output — not the Learn step's note-taking, which is what
        # "first task with an evidence cell" was picking up. Capstone days
        # have no Work Mission label, so their own first task's evidence is
        # the artefact ("Project brief", "Working version", "Final portfolio
        # package"). The ROADMAP's Daily Deliverable column is not used: it
        # is the same boilerplate sentence on 79 of the 90 days.
        deliverable = (
            evidence_for("work mission", "portfolio", "capstone")
            or first_work_task(sheet)["evidence"]
            or next((t["evidence"] for t in sheet["tasks"] if t["evidence"]), "")
            or "Working file, screenshot or recorded result"
        )

        # Days 80-90 all share one source lesson ("Module 20 — Capstone /
        # Portfolio"), so the lesson name would repeat 11 times in the task
        # list. Those sheets name their tasks for the artefact being built
        # ("Capstone Brief", "Build — Part 1", "Present"), which is the real
        # distinguishing title the workbook already provides.
        if is_capstone:
            lesson_title = first_work_task(sheet)["title"] or sheet["sourceLessons"]
        else:
            lesson_title = sheet["sourceLessons"] or meta["lesson"] or f"Day {day}"
        source_ref = (
            f"{sheet['sourceLessons']} (source page {sheet['sourcePages']})"
            if sheet["sourcePages"]
            else sheet["sourceLessons"]
        )

        workstreams[key]["tasks"].append(
            {
                "taskId": f"day-{day:02d}",
                "dayNumber": day,
                "title": lesson_title,
                "tier": "ASSESSMENT" if is_capstone else "REQUIRED",
                "estimatedMinutes": total,
                "whatYoureDoing": learn or mission or sheet["learningFocus"],
                "whyItMatters": (
                    f"Builds your {skill.lower()} skill — part of "
                    f"{stage}, day {day} of 90."
                ),
                "skillDeveloped": skill,
                "understandText": source_ref or lesson_title,
                "watchLabel": None,
                "practiseText": practice,
                "assignmentText": mission or practice or learn,
                "deliverableLabel": deliverable,
                # The workbook titles its days after the source lessons
                # ("Lesson 1A - ..."), which is the join to lessons.json.
                "lessonCodes": re.findall(r"Lesson\s+(\d+[A-Z])", lesson_title),
                "subSteps": sub_steps,
                "assessmentCriteria": criteria_for(sub_steps, deliverable),
                "reviewQuestions": sheet["questions"],
            }
        )

    package = {
        "programmeId": "digital-foundation",
        "generatedFrom": WORKBOOK.name,
        "totalDays": len(roadmap),
        "workstreams": list(workstreams.values()),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(package, indent=2, ensure_ascii=False) + "\n")

    tasks = sum(len(w["tasks"]) for w in package["workstreams"])
    steps = sum(len(t["subSteps"]) for w in package["workstreams"] for t in w["tasks"])
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  workstreams : {len(package['workstreams'])}")
    print(f"  tasks(days) : {tasks}")
    print(f"  sub-steps   : {steps}")
    print(f"  size        : {OUT.stat().st_size / 1024:.0f} KB")
    for stage_label, stage_id in STAGE_IDS.items():
        ws = [w for w in package["workstreams"] if w["stageId"] == stage_id]
        n = sum(len(w["tasks"]) for w in ws)
        print(f"  {stage_id} {stage_label:24} {len(ws):2} workstreams, {n:2} days")


if __name__ == "__main__":
    build()
